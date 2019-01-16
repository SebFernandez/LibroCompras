import openpyxl

maxRow = 0
iva = ['10.50', '21', '27']
tipoComprobante = {'FCC':{'A':'001', 'B': '006', 'C' : '011'},
                'NCC':{'A':'003', 'B': '008', 'C' : '013'},
                'NDC':{'A':'002', 'B': '007', 'C' : '012'},
                'REC':{'A':'004', 'B': '009', 'C' : '015'}
                }

compra = openpyxl.load_workbook ('Compras.xlsx')
model  = openpyxl.load_workbook ('MODELO DE FORMULARIO.xlsx')

#Hoja de COMPRAS
sheetCompra = compra.active

#Hoja del MODELO
sheetModel = model.get_sheet_by_name('Carga de Datos')

def codigoDeComprobante (tipo, varToTrim):
    if (tipo in tipoComprobante):
        if (varToTrim [:1] in tipoComprobante [tipo]):
            return tipoComprobante [tipo][varToTrim [:1]]
        elif (varToTrim [:1] is '0'):
            return tipoComprobante [tipo]['B']
        elif (tipo is 'REC'):
            return tipoComprobante [tipo]['C']  
        else:
            return '-----'

def ptoDeVenta (varToTrim):
    return varToTrim [1:5].split("-")[0]

def nroComprobante (varToTrim):
    varAux = varToTrim.split("-")[1]
    varAux2 = varAux.split ("/")[0]
    return varAux2

def cuit (idComprobante):
    auxString = idComprobante [:2] + idComprobante [3:11] + idComprobante [12:13]
    return auxString

def calculoIVA (neto, impValor):
    varNeto = float (neto)
    varImpValor = float (impValor)

    varResult = (varImpValor/varNeto) * 100

    if (varResult == 21 or varResult == 27):
        varResult = int (varResult)
    else:
        varResult = round (varResult,1)

    return str (varResult)

#Se resta 8 en el for porque la planilla de Compras tiene 8 filas semi-completas que no se usan.
for i in range (sheetCompra.max_row-8):
    #Copia fecha
    sheetModel.cell (row = 13+i, column = 1).value = sheetCompra.cell (row=2+i, column = 1).value
    aux = str (sheetCompra.cell (row=2+i, column = 2).value)
    aux2 = str (sheetCompra.cell (row=2+i, column = 3).value)
    aux3 = str (sheetCompra.cell (row=2+i, column = 4).value)
    aux4 = str (sheetCompra.cell (row=2+i, column = 7).value)
    aux5 = str (sheetCompra.cell (row=2+i, column = 8).value)

    #Copia comprobante, pto venta, nro comprobante, cuit, etc.
    sheetModel.cell (row = 13+i, column = 2).value = codigoDeComprobante (aux,aux2)
    sheetModel.cell (row = 13+i, column = 3).value = ptoDeVenta (aux2)
    sheetModel.cell (row = 13+i, column = 4).value = nroComprobante (aux2)
    sheetModel.cell (row = 13+i, column = 7).value = cuit (aux3)
    sheetModel.cell (row = 13+i, column = 8).value = sheetCompra.cell (row=2+i, column = 6).value

    #Copia valores de conceptos sobre importes; ergo, nros contables.
    #Ordenado en cómo avanza de izq. a der. las columnas de COMPRAS.
    sheetModel.cell (row = 13+i, column = 20).value = sheetCompra.cell (row=2+i, column = 7).value
    sheetModel.cell (row = 13+i, column = 12).value = sheetCompra.cell (row=2+i, column = 9).value
    sheetModel.cell (row = 13+i, column = 14).value = sheetCompra.cell (row=2+i, column = 10).value
    sheetModel.cell (row = 13+i, column = 10).value = sheetCompra.cell (row=2+i, column = 11).value
    sheetModel.cell (row = 13+i, column = 9).value = sheetCompra.cell (row=2+i, column = 12).value
    try:
        sheetModel.cell (row = 13+i, column = 21).value = calculoIVA (aux4,aux5)
    except:
        pass
    
model.save ('MODELO DE FORMULARIO.xlsx')